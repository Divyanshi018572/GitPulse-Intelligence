import io
import json
import urllib.parse
import asyncio
from fastapi import APIRouter, HTTPException, Header
from fastapi.responses import StreamingResponse

from core.github import get_github_data, GITHUB_API
from core.utils import validate_username

router = APIRouter()

@router.get("/api/export")
async def export_excel(users: str = "[]"):
    try:
        raw = urllib.parse.unquote(users)
        parsed = json.loads(raw)
        if not isinstance(parsed, list):
            raise ValueError("Expected a JSON array.")
    except (json.JSONDecodeError, ValueError) as e:
        raise HTTPException(status_code=400, detail=f"Invalid users data: {e}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl not installed.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Talent Report"

    headers = ["Name", "Username", "Bio", "Location", "Company", "Public Repos", "Followers", "URL"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for u in parsed:
        ws.append([
            u.get("name") or "",
            u.get("login") or "",
            u.get("bio") or "",
            u.get("location") or "",
            u.get("company") or "",
            u.get("public_repos") or 0,
            u.get("followers") or 0,
            u.get("html_url") or "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        headers={"Content-Disposition": 'attachment; filename="talent_report.xlsx"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@router.get("/api/export_pdf/{username}")
async def export_pdf(username: str, x_github_token: str = Header(None)):
    validate_username(username)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    except ImportError:
        raise HTTPException(status_code=503, detail="reportlab not installed.")

    user_res, repos_res = await asyncio.gather(
        get_github_data(f"{GITHUB_API}/users/{username}", user_token=x_github_token),
        get_github_data(f"{GITHUB_API}/users/{username}/repos?per_page=6&sort=stars", user_token=x_github_token),
    )
    if user_res.status_code != 200:
        raise HTTPException(status_code=404, detail="User not found.")

    u     = user_res.json()
    repos = repos_res.json() if repos_res.status_code == 200 else []

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    accent = colors.HexColor("#1f6feb")
    dark   = colors.HexColor("#0d1117")
    muted  = colors.HexColor("#57606a")

    h1 = ParagraphStyle("h1", fontSize=22, fontName="Helvetica-Bold", textColor=dark, spaceAfter=4)
    h2 = ParagraphStyle("h2", fontSize=13, fontName="Helvetica-Bold", textColor=accent, spaceBefore=14, spaceAfter=6)
    body = ParagraphStyle("body", fontSize=10, fontName="Helvetica", textColor=dark, spaceAfter=4, leading=15)
    meta = ParagraphStyle("meta", fontSize=9, fontName="Helvetica", textColor=muted, spaceAfter=3)
    label = ParagraphStyle("label", fontSize=8, fontName="Helvetica-Bold", textColor=muted, spaceAfter=2, letterSpacing=1)

    story.append(Paragraph(u.get("name") or u.get("login"), h1))
    story.append(Paragraph(f'@{u.get("login")}  ·  github.com/{u.get("login")}', meta))
    if u.get("bio"):
        story.append(Paragraph(u["bio"], body))
    story.append(Spacer(1, 6))

    stats = [
        ["📁 Public Repos", "👥 Followers", "👤 Following"],
        [str(u.get("public_repos",0)), str(u.get("followers",0)), str(u.get("following",0))],
    ]
    t = Table(stats, colWidths=[5*cm, 5*cm, 5*cm])
    t.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (-1,0), "Helvetica"), ("FONTSIZE",  (0,0), (-1,0), 8), ("TEXTCOLOR", (0,0), (-1,0), muted),
        ("FONTNAME",  (0,1), (-1,1), "Helvetica-Bold"), ("FONTSIZE",  (0,1), (-1,1), 16), ("TEXTCOLOR", (0,1), (-1,1), accent),
        ("ALIGN",     (0,0), (-1,-1), "CENTER"), ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#d0d7de"), spaceAfter=10))

    story.append(Paragraph("PROFILE DETAILS", label))
    details = []
    if u.get("location"):  details.append(f"📍 {u['location']}")
    if u.get("company"):   details.append(f"🏢 {u['company']}")
    if u.get("blog"):      details.append(f"🔗 {u['blog']}")
    if u.get("email"):     details.append(f"✉️  {u['email']}")
    for d in details:
        story.append(Paragraph(d, body))
    story.append(Spacer(1, 6))

    if repos:
        story.append(Paragraph("TOP REPOSITORIES", label))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#d0d7de"), spaceAfter=6))
        for r in repos[:6]:
            story.append(Paragraph(f'<b>{r.get("name","")}</b>', body))
            if r.get("description"):
                story.append(Paragraph(r["description"], meta))
            row_meta = f'⭐ {r.get("stargazers_count",0)}  ·  🍴 {r.get("forks_count",0)}'
            if r.get("language"):
                row_meta += f'  ·  💻 {r["language"]}'
            story.append(Paragraph(row_meta, meta))
            story.append(Spacer(1, 4))

    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#d0d7de"), spaceBefore=10, spaceAfter=6))
    story.append(Paragraph("Generated by GitHub Talent Finder", meta))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, headers={"Content-Disposition": f'attachment; filename="{username}_profile.pdf"'}, media_type="application/pdf")

@router.get("/api/export_shortlist_pdf")
async def export_shortlist_pdf(candidates: str = "[]"):
    try:
        raw       = urllib.parse.unquote(candidates)
        parsed    = json.loads(raw)
        if not isinstance(parsed, list):
            raise ValueError
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid candidates data.")

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    except ImportError:
        raise HTTPException(status_code=503, detail="reportlab not installed.")

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    accent  = colors.HexColor("#1f6feb")
    dark    = colors.HexColor("#0d1117")
    muted   = colors.HexColor("#57606a")
    green   = colors.HexColor("#1a7f37")

    h1    = ParagraphStyle("h1",  fontSize=20, fontName="Helvetica-Bold", textColor=dark,   spaceAfter=4)
    h2    = ParagraphStyle("h2",  fontSize=13, fontName="Helvetica-Bold", textColor=accent, spaceBefore=16, spaceAfter=6)
    body  = ParagraphStyle("body",fontSize=10, fontName="Helvetica",      textColor=dark,   spaceAfter=3,   leading=14)
    meta  = ParagraphStyle("meta",fontSize=9,  fontName="Helvetica",      textColor=muted,  spaceAfter=2)
    note  = ParagraphStyle("note",fontSize=10, fontName="Helvetica-Oblique", textColor=green, spaceAfter=4, leading=14)
    label = ParagraphStyle("lbl", fontSize=8,  fontName="Helvetica-Bold", textColor=muted,  spaceAfter=3, letterSpacing=1)

    story = []
    story.append(Paragraph("Candidate Shortlist Report", h1))
    story.append(Paragraph(f"Generated by GitHub Talent Finder · {len(parsed)} candidates", meta))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#d0d7de"), spaceBefore=8, spaceAfter=12))

    for i, c in enumerate(parsed):
        story.append(Paragraph(f"{i+1}. {c.get('name') or c.get('login','')}", h2))
        story.append(Paragraph(f"@{c.get('login','')}  ·  github.com/{c.get('login','')}", meta))
        if c.get("bio"):
            story.append(Paragraph(c["bio"], body))

        details = []
        if c.get("location"):  details.append(f"📍 {c['location']}")
        if c.get("public_repos") is not None: details.append(f"📁 {c['public_repos']} repos")
        if c.get("followers")   is not None: details.append(f"👥 {c['followers']} followers")
        if details:
            story.append(Paragraph("  ·  ".join(details), meta))

        if c.get("note"):
            story.append(Spacer(1, 4))
            story.append(Paragraph("RECRUITER NOTE", label))
            story.append(Paragraph(c["note"], note))

        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#d0d7de"), spaceBefore=8, spaceAfter=4))

    story.append(Paragraph("Generated by GitHub Talent Finder", meta))
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, headers={"Content-Disposition": 'attachment; filename="shortlist_report.pdf"'}, media_type="application/pdf")
